from flask import Flask, request, jsonify, send_file, session, redirect, url_for
from flask_cors import CORS
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import os
from werkzeug.utils import secure_filename
from werkzeug.security import check_password_hash, generate_password_hash
from datetime import datetime
import secrets

app = Flask(__name__)
app.secret_key = secrets.token_hex(32)  # Genera chiave segreta random
CORS(app)

# ⚠️ CONFIGURAZIONE SICUREZZA - CAMBIA QUESTI VALORI!
USERS = {
    'admin': generate_password_hash('BilancioMVP2024!')  # Username: admin, Password: BilancioMVP2024!
}

# Configurazione
UPLOAD_FOLDER = '/tmp/uploads'
OUTPUT_FOLDER = '/tmp/outputs'
ALLOWED_EXTENSIONS = {'pdf', 'xlsx', 'xls'}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def login_required(f):
    """Decorator per proteggere le route"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def get_macro_cluster_iv_direttiva(codice):
    """Mapping verso macro cluster IV Direttiva CEE"""
    first_part = codice.split('/')[0]
    first_digits = int(first_part)
    
    if first_digits <= 54:
        if first_part in ['03', '04', '06', '07', '09']:
            return "B) Immobilizzazioni"
        elif first_part in ['11', '15', '18', '24']:
            return "C) Attivo Circolante"
        elif first_part == '26':
            return "D) Ratei e Risconti"
        elif first_part in ['28']:
            return "A) Patrimonio Netto"
        elif first_part in ['31', '38', '40', '41', '46', '48', '50', '52']:
            return "D) Debiti"
        elif first_part == '54':
            return "E) Ratei e Risconti (Passivo)"
        else:
            return "Altro SP"
    else:
        if first_part in ['58', '64', '80']:
            return "A) Valore della Produzione"
        elif first_part in ['66', '68', '70', '72', '74', '75', '84']:
            return "B) Costi della Produzione"
        elif first_part in ['87', '88']:
            return "C) Proventi e Oneri Finanziari"
        elif first_part == '96':
            return "Imposte"
        else:
            return "Altro CE"

def extract_bilancio_from_pdf(pdf_path):
    """Estrae dati dal PDF del bilancio"""
    data = []
    
    with pdfplumber.open(pdf_path) as pdf:
        num_pages = len(pdf.pages)
        pages_to_process = max(1, num_pages - 2)
        
        for page_num in range(pages_to_process):
            page = pdf.pages[page_num]
            tables = page.extract_tables()
            
            if not tables:
                continue
            
            for table in tables:
                for row in table:
                    if not row or len(row) < 2:
                        continue
                    
                    if 'Conto' in str(row[0]) or 'Descrizione' in str(row[0]):
                        continue
                    
                    if not row[0] or row[0].strip() == '':
                        continue
                    
                    first_col = row[0].strip()
                    
                    match_normal = re.match(r'^(\d{2}/\d{4}/\d{4})\s+(.+)$', first_col)
                    match_special = re.match(r'^(\d{2}/\*{8})\s+(.+)$', first_col)
                    
                    if match_normal and '****' not in first_col:
                        codice = match_normal.group(1)
                        descrizione = match_normal.group(2).strip()
                    elif match_special:
                        codice = match_special.group(1)
                        descrizione = match_special.group(2).strip()
                    else:
                        continue
                    
                    attivita = row[1].strip() if len(row) > 1 and row[1] else ''
                    passivita = row[2].strip() if len(row) > 2 and row[2] else ''
                    
                    def parse_value(val_str):
                        if not val_str or val_str == '':
                            return None
                        val_str = val_str.replace('.', '').replace(',', '.')
                        try:
                            return float(val_str)
                        except:
                            return None
                    
                    val_attivita = parse_value(attivita)
                    val_passivita = parse_value(passivita)
                    
                    if val_attivita is not None and val_passivita is None:
                        raw_amount = val_attivita
                        colonna = 'Attività/Costi'
                    elif val_passivita is not None and val_attivita is None:
                        raw_amount = val_passivita
                        colonna = 'Passività/Ricavi'
                    elif val_attivita is not None and val_passivita is not None:
                        raw_amount = val_attivita if val_attivita != 0 else val_passivita
                        colonna = 'Attività/Costi' if val_attivita != 0 else 'Passività/Ricavi'
                    else:
                        continue
                    
                    first_digits = int(codice.split('/')[0])
                    
                    if first_digits <= 54:
                        tipo = "Stato Patrimoniale"
                        amount = raw_amount if colonna == 'Attività/Costi' else -raw_amount
                    else:
                        tipo = "Conto Economico"
                        amount = raw_amount if colonna == 'Attività/Costi' else -raw_amount
                    
                    macro_cluster = get_macro_cluster_iv_direttiva(codice)
                    
                    data.append({
                        'Codice Conto': codice,
                        'Descrizione': descrizione,
                        'Tipo': tipo,
                        'Amount': amount,
                        'Macro Cluster IV Direttiva': macro_cluster
                    })
    
    return pd.DataFrame(data)

def create_excel_output(df, output_path):
    """Crea Excel con bilancino pulito"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bilancino Pulito"
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['Codice Conto', 'Descrizione', 'Tipo', 'Amount', 'Macro Cluster IV Direttiva']
    ws.append(headers)
    
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = border
            if cell.column == 4:
                cell.number_format = '#,##0.00'
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 35
    
    wb.save(output_path)

@app.route('/')
def home():
    if 'logged_in' not in session:
        return redirect(url_for('login'))
    return redirect(url_for('dashboard'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        
        if username in USERS and check_password_hash(USERS[username], password):
            session['logged_in'] = True
            session['username'] = username
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'error': 'Credenziali non valide'})
    
    return '''
    <!DOCTYPE html>
    <html lang="it">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Login - Bilancio Processor</title>
        <style>
            * { margin: 0; padding: 0; box-sizing: border-box; }
            body {
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                min-height: 100vh;
                display: flex;
                align-items: center;
                justify-content: center;
                padding: 20px;
            }
            .login-container {
                background: white;
                border-radius: 20px;
                box-shadow: 0 20px 60px rgba(0,0,0,0.3);
                padding: 50px 40px;
                max-width: 400px;
                width: 100%;
                text-align: center;
            }
            .logo {
                font-size: 64px;
                margin-bottom: 20px;
            }
            h1 {
                color: #333;
                margin-bottom: 10px;
                font-size: 28px;
            }
            .subtitle {
                color: #666;
                margin-bottom: 30px;
                font-size: 14px;
            }
            .lock-badge {
                display: inline-block;
                background: #ffc107;
                color: #333;
                padding: 5px 15px;
                border-radius: 20px;
                font-size: 12px;
                font-weight: 600;
                margin-bottom: 30px;
            }
            .input-group {
                margin-bottom: 20px;
                text-align: left;
            }
            label {
                display: block;
                color: #666;
                font-size: 14px;
                margin-bottom: 8px;
                font-weight: 600;
            }
            input {
                width: 100%;
                padding: 15px;
                border: 2px solid #e0e0e0;
                border-radius: 10px;
                font-size: 16px;
                transition: border-color 0.3s;
            }
            input:focus {
                outline: none;
                border-color: #667eea;
            }
            .login-btn {
                width: 100%;
                padding: 15px;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
                border: none;
                border-radius: 10px;
                font-size: 16px;
                font-weight: 600;
                cursor: pointer;
                transition: transform 0.2s;
                margin-top: 20px;
            }
            .login-btn:hover {
                transform: translateY(-2px);
                box-shadow: 0 10px 20px rgba(102, 126, 234, 0.3);
            }
            .login-btn:disabled {
                opacity: 0.6;
                cursor: not-allowed;
            }
            .error {
                background: #ffe0e0;
                color: #c62828;
                padding: 12px;
                border-radius: 8px;
                margin-top: 15px;
                font-size: 14px;
                display: none;
            }
            .spinner {
                display: none;
                width: 20px;
                height: 20px;
                border: 3px solid rgba(255,255,255,0.3);
                border-top-color: white;
                border-radius: 50%;
                animation: spin 0.8s linear infinite;
                margin: 0 auto;
            }
            @keyframes spin {
                to { transform: rotate(360deg); }
            }
        </style>
    </head>
    <body>
        <div class="login-container">
            <div class="logo">🔒</div>
            <h1>Accesso Riservato</h1>
            <p class="subtitle">Bilancio Processor MVP</p>
            <span class="lock-badge">🛡️ AREA PRIVATA</span>
            
            <form id="loginForm">
                <div class="input-group">
                    <label for="username">Username</label>
                    <input type="text" id="username" name="username" required autocomplete="username">
                </div>
                <div class="input-group">
                    <label for="password">Password</label>
                    <input type="password" id="password" name="password" required autocomplete="current-password">
                </div>
                <button type="submit" class="login-btn" id="loginBtn">
                    <span id="btnText">Accedi</span>
                    <div class="spinner" id="spinner"></div>
                </button>
                <div class="error" id="error"></div>
            </form>
        </div>
        
        <script>
            document.getElementById('loginForm').addEventListener('submit', async (e) => {
                e.preventDefault();
                
                const username = document.getElementById('username').value;
                const password = document.getElementById('password').value;
                const loginBtn = document.getElementById('loginBtn');
                const btnText = document.getElementById('btnText');
                const spinner = document.getElementById('spinner');
                const error = document.getElementById('error');
                
                // Disabilita bottone
                loginBtn.disabled = true;
                btnText.style.display = 'none';
                spinner.style.display = 'block';
                error.style.display = 'none';
                
                try {
                    const response = await fetch('/login', {
                        method: 'POST',
                        headers: { 'Content-Type': 'application/json' },
                        body: JSON.stringify({ username, password })
                    });
                    
                    const data = await response.json();
                    
                    if (data.success) {
                        window.location.href = '/dashboard';
                    } else {
                        error.textContent = '❌ ' + data.error;
                        error.style.display = 'block';
                        loginBtn.disabled = false;
                        btnText.style.display = 'block';
                        spinner.style.display = 'none';
                    }
                } catch (err) {
                    error.textContent = '❌ Errore di connessione';
                    error.style.display = 'block';
                    loginBtn.disabled = false;
                    btnText.style.display = 'block';
                    spinner.style.display = 'none';
                }
            });
        </script>
    </body>
    </html>
    '''

@app.route('/dashboard')
@login_required
def dashboard():
    return '''
    <!DOCTYPE html>
    <html lang="it">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Dashboard - Bilancio Processor</title>
        <style>
            * { margin: 0; padding: 0; box-sizing: border-box; }
            body {
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                min-height: 100vh;
                padding: 20px;
            }
            .header {
                background: rgba(255,255,255,0.95);
                padding: 15px 30px;
                border-radius: 15px;
                margin-bottom: 20px;
                display: flex;
                justify-content: space-between;
                align-items: center;
                box-shadow: 0 5px 20px rgba(0,0,0,0.1);
            }
            .header-left {
                display: flex;
                align-items: center;
                gap: 15px;
            }
            .logo-small {
                font-size: 32px;
            }
            .header h2 {
                color: #333;
                font-size: 20px;
            }
            .user-badge {
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
                padding: 8px 20px;
                border-radius: 20px;
                font-size: 14px;
                font-weight: 600;
            }
            .logout-btn {
                background: #f44336;
                color: white;
                border: none;
                padding: 8px 20px;
                border-radius: 8px;
                cursor: pointer;
                font-weight: 600;
                margin-left: 15px;
                transition: background 0.3s;
            }
            .logout-btn:hover {
                background: #d32f2f;
            }
            .container {
                background: white;
                border-radius: 20px;
                box-shadow: 0 20px 60px rgba(0,0,0,0.3);
                padding: 40px;
                max-width: 900px;
                margin: 0 auto;
            }
            h1 { color: #333; margin-bottom: 30px; font-size: 28px; }
            .upload-area {
                border: 3px dashed #667eea;
                border-radius: 15px;
                padding: 60px 40px;
                text-align: center;
                cursor: pointer;
                transition: all 0.3s;
                background: #f8f9ff;
            }
            .upload-area:hover {
                background: #f0f2ff;
                border-color: #764ba2;
                transform: scale(1.02);
            }
            .upload-icon { font-size: 64px; margin-bottom: 20px; }
            .upload-text { color: #667eea; font-size: 18px; font-weight: 600; margin-bottom: 10px; }
            .upload-hint { color: #999; font-size: 14px; }
            #fileInput { display: none; }
            .file-info {
                background: #f8f9ff;
                padding: 20px;
                border-radius: 10px;
                margin: 20px 0;
                display: none;
            }
            .file-name { font-weight: 600; color: #333; margin-bottom: 10px; }
            .progress-bar {
                width: 100%;
                height: 8px;
                background: #e0e0e0;
                border-radius: 4px;
                overflow: hidden;
                margin: 15px 0;
            }
            .progress-fill {
                height: 100%;
                background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
                width: 0%;
                transition: width 0.3s;
            }
            .status { color: #666; font-size: 14px; margin-top: 10px; }
            .results {
                display: none;
                margin-top: 30px;
                padding: 20px;
                background: #f8f9ff;
                border-radius: 10px;
            }
            .results h3 { color: #333; margin-bottom: 15px; }
            .stat {
                display: flex;
                justify-content: space-between;
                padding: 10px 0;
                border-bottom: 1px solid #e0e0e0;
            }
            .stat:last-child { border-bottom: none; }
            .stat-label { color: #666; }
            .stat-value { font-weight: 600; color: #667eea; }
            .download-btn {
                display: inline-block;
                margin-top: 20px;
                padding: 15px 40px;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
                text-decoration: none;
                border-radius: 10px;
                font-weight: 600;
                transition: transform 0.2s;
                border: none;
                cursor: pointer;
                font-size: 16px;
            }
            .download-btn:hover {
                transform: translateY(-2px);
                box-shadow: 0 10px 20px rgba(102, 126, 234, 0.3);
            }
            .error {
                background: #ffe0e0;
                color: #d32f2f;
                padding: 15px;
                border-radius: 10px;
                margin-top: 20px;
                display: none;
            }
        </style>
    </head>
    <body>
        <div class="header">
            <div class="header-left">
                <div class="logo-small">🚀</div>
                <h2>Bilancio Processor MVP</h2>
            </div>
            <div>
                <span class="user-badge">👤 Utente connesso</span>
                <button class="logout-btn" onclick="logout()">Esci</button>
            </div>
        </div>
        
        <div class="container">
            <h1>Carica il tuo bilancino di verifica</h1>
            
            <div class="upload-area" id="uploadArea">
                <div class="upload-icon">📄</div>
                <div class="upload-text">Trascina qui il tuo file o clicca per selezionarlo</div>
                <div class="upload-hint">PDF o Excel • Max 10MB</div>
            </div>
            
            <input type="file" id="fileInput" accept=".pdf,.xlsx,.xls">
            
            <div class="file-info" id="fileInfo">
                <div class="file-name" id="fileName"></div>
                <div class="progress-bar">
                    <div class="progress-fill" id="progressFill"></div>
                </div>
                <div class="status" id="status">Caricamento in corso...</div>
            </div>
            
            <div class="error" id="error"></div>
            
            <div class="results" id="results">
                <h3>✅ Elaborazione completata!</h3>
                <div class="stat">
                    <span class="stat-label">Conti estratti:</span>
                    <span class="stat-value" id="totalConti">-</span>
                </div>
                <div class="stat">
                    <span class="stat-label">Stato Patrimoniale:</span>
                    <span class="stat-value" id="contiSP">-</span>
                </div>
                <div class="stat">
                    <span class="stat-label">Conto Economico:</span>
                    <span class="stat-value" id="contiCE">-</span>
                </div>
                <div class="stat">
                    <span class="stat-label">Quadratura:</span>
                    <span class="stat-value" id="quadratura">-</span>
                </div>
                <a href="#" class="download-btn" id="downloadBtn">📥 Scarica Excel Elaborato</a>
            </div>
        </div>
        
        <script>
            const uploadArea = document.getElementById('uploadArea');
            const fileInput = document.getElementById('fileInput');
            const fileInfo = document.getElementById('fileInfo');
            const fileName = document.getElementById('fileName');
            const progressFill = document.getElementById('progressFill');
            const status = document.getElementById('status');
            const results = document.getElementById('results');
            const error = document.getElementById('error');
            
            uploadArea.addEventListener('click', () => fileInput.click());
            
            uploadArea.addEventListener('dragover', (e) => {
                e.preventDefault();
                uploadArea.classList.add('dragover');
            });
            
            uploadArea.addEventListener('dragleave', () => {
                uploadArea.classList.remove('dragover');
            });
            
            uploadArea.addEventListener('drop', (e) => {
                e.preventDefault();
                uploadArea.classList.remove('dragover');
                const files = e.dataTransfer.files;
                if (files.length > 0) {
                    handleFile(files[0]);
                }
            });
            
            fileInput.addEventListener('change', (e) => {
                if (e.target.files.length > 0) {
                    handleFile(e.target.files[0]);
                }
            });
            
            function handleFile(file) {
                error.style.display = 'none';
                results.style.display = 'none';
                fileInfo.style.display = 'block';
                fileName.textContent = file.name;
                progressFill.style.width = '30%';
                
                const formData = new FormData();
                formData.append('file', file);
                
                fetch('/upload', {
                    method: 'POST',
                    body: formData
                })
                .then(response => {
                    progressFill.style.width = '70%';
                    return response.json();
                })
                .then(data => {
                    progressFill.style.width = '100%';
                    if (data.success) {
                        setTimeout(() => showResults(data), 500);
                    } else {
                        showError(data.error);
                    }
                })
                .catch(err => {
                    showError('Errore durante l\\'elaborazione: ' + err.message);
                });
            }
            
            function showResults(data) {
                fileInfo.style.display = 'none';
                results.style.display = 'block';
                
                document.getElementById('totalConti').textContent = data.stats.total_conti;
                document.getElementById('contiSP').textContent = data.stats.conti_sp;
                document.getElementById('contiCE').textContent = data.stats.conti_ce;
                document.getElementById('quadratura').textContent = 
                    data.stats.quadra ? '✅ Perfetto (0,00 €)' : '⚠️ ' + data.stats.totale.toFixed(2) + ' €';
                
                document.getElementById('downloadBtn').href = '/download/' + data.file_id;
            }
            
            function showError(message) {
                fileInfo.style.display = 'none';
                error.style.display = 'block';
                error.textContent = '❌ ' + message;
            }
            
            function logout() {
                if (confirm('Vuoi davvero uscire?')) {
                    window.location.href = '/logout';
                }
            }
        </script>
    </body>
    </html>
    '''

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/upload', methods=['POST'])
@login_required
def upload_file():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'Nessun file caricato'})
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'success': False, 'error': 'Nessun file selezionato'})
    
    if not allowed_file(file.filename):
        return jsonify({'success': False, 'error': 'Formato file non supportato'})
    
    try:
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        file_id = f"{timestamp}_{filename}"
        input_path = os.path.join(UPLOAD_FOLDER, file_id)
        file.save(input_path)
        
        ext = filename.rsplit('.', 1)[1].lower()
        if ext == 'pdf':
            df = extract_bilancio_from_pdf(input_path)
        else:
            return jsonify({'success': False, 'error': 'Formato Excel non ancora supportato'})
        
        if df.empty:
            return jsonify({'success': False, 'error': 'Impossibile estrarre dati dal file'})
        
        output_filename = f"elaborato_{file_id.replace('.pdf', '.xlsx')}"
        output_path = os.path.join(OUTPUT_FOLDER, output_filename)
        create_excel_output(df, output_path)
        
        df_sp = df[df['Tipo'] == 'Stato Patrimoniale']
        df_ce = df[df['Tipo'] == 'Conto Economico']
        totale = df['Amount'].sum()
        
        stats = {
            'total_conti': len(df),
            'conti_sp': len(df_sp),
            'conti_ce': len(df_ce),
            'totale': float(totale),
            'quadra': abs(totale) < 0.10
        }
        
        return jsonify({
            'success': True,
            'file_id': output_filename,
            'stats': stats
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/download/<file_id>')
@login_required
def download_file(file_id):
    try:
        file_path = os.path.join(OUTPUT_FOLDER, file_id)
        return send_file(file_path, as_attachment=True, download_name=f"bilancio_elaborato.xlsx")
    except Exception as e:
        return jsonify({'error': str(e)}), 404

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
